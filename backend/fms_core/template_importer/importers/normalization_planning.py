from fms_core.template_importer.row_handlers.normalization_planning import NormalizationPlanningRowHandler
from fms_core.template_prefiller._utils import load_position_dict
from fms_core.templates import NORMALIZATION_PLANNING_TEMPLATE, NORMALIZATION_TEMPLATE
from fms_core.models import Container
from fms_core.utils import str_cast_and_normalize

from openpyxl.reader.excel import load_workbook
from django.conf import settings

from ...containers import CONTAINER_KIND_SPECS
from .._utils import float_to_decimal_and_none
from ._generic import GenericImporter

import zipfile
import io
import os
from datetime import datetime
from typing import Union

class NormalizationPlanningImporter(GenericImporter):
    """
         Template importer for the Normalization Protocol.

         Args:
             `sheet`: The template to ingest.

         Returns:
             A detailed validation of the data trying to be ingested as a result.
    """

    SHEETS_INFO = NORMALIZATION_PLANNING_TEMPLATE["sheets info"]

    def __init__(self):
        super().__init__()
        self.initialize_data_for_template()

    def initialize_data_for_template(self):
        # Maybe useful ?
        pass

    def import_template_inner(self):
        sheet = self.sheets['Normalization']

        mapping_rows_template = []
        # For each row initialize the object that is going to be prefilled in the normalization template
        for row_id, row_data in enumerate(sheet.rows):
            source_sample = {
                'name': str_cast_and_normalize(row_data['Sample Name']),
                'container': {'barcode': str_cast_and_normalize(row_data['Source Container Barcode'])},
                'coordinates': str_cast_and_normalize(row_data['Source Container Coord']),
            }

            destination_sample = {
                'coordinates': str_cast_and_normalize(row_data['Destination Container Coord']),
                'container': {
                    'barcode': str_cast_and_normalize(row_data['Destination Container Barcode']),
                    'name': str_cast_and_normalize(row_data['Destination Container Name']),
                    'kind': str_cast_and_normalize(row_data['Destination Container Kind']),
                    'coordinates': str_cast_and_normalize(row_data['Destination Parent Container Coord']),
                    'parent_barcode': str_cast_and_normalize(row_data['Destination Parent Container Barcode']),
                },
            }

            measurements = {
                'volume': float_to_decimal_and_none(row_data['Final Volume (uL)']),
                'na_quantity': float_to_decimal_and_none(row_data['NA Quantity (ng)']),
                'concentration_ngul': float_to_decimal_and_none(row_data['Norm. Conc. (ng/uL)']),
                'concentration_nm': float_to_decimal_and_none(row_data['Norm. Conc. (nM)']),
            }

            normalization_kwargs = dict(
                source_sample=source_sample,
                destination_sample=destination_sample,
                measurements=measurements
            )

            (result, row_mapping) = self.handle_row(
                row_handler_class=NormalizationPlanningRowHandler,
                sheet=sheet,
                row_i=row_id,
                **normalization_kwargs,
            )

            mapping_rows_template.append(row_mapping)

        if not self.dry_run:
            # Populate files
            
            # Create robot file and complete mapping_rows_template with the 
            robot_filename = self.prepare_robot_file(mapping_rows_template)

            if robot_filename is not None:
                # TODO: implement independent functions / services for this
                # TODO: a lot
                #Populate template
                out_stream = io.BytesIO()

                filename = "/".join(NORMALIZATION_TEMPLATE["identity"]["file"].split("/")[-2:])
                template_path = os.path.join(settings.STATIC_ROOT, filename)
                workbook = load_workbook(filename=template_path)
                position_dict = load_position_dict(workbook, NORMALIZATION_TEMPLATE["sheets info"], NORMALIZATION_TEMPLATE["prefill info"])

                try:
                    for sheet_name, sheet_dict in position_dict.items():
                        current_sheet = workbook[sheet_name]
                        for i, entry in enumerate(mapping_rows_template):
                            for sheet in NORMALIZATION_TEMPLATE["sheets info"]:
                                for header_index, template_column in enumerate(sheet['headers']):
                                    if sheet["name"] == sheet_name:
                                        current_sheet.cell(row=sheet_dict["header_offset"] + i, column=header_index + 1).value = entry[template_column]
                    workbook.save(out_stream)
                except Exception as e:
                    print("Failed to fill result template : " + str(e))

                # TODO: Random name from id generator
                output_zip_name = f"Normalization_planning_output_{datetime.today().strftime('%Y-%m-%d')}"
                normalization_template_filename = filename.split('/')[1]

                # Zip files
                try:
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                        # Add normalization prefilled template
                        file = out_stream.getvalue()
                        zip_file.writestr(output_zip_name + '/' + normalization_template_filename, file)
                        # TODO: add robot csv file
                except Exception as e:
                    print("Failed to zip the result file: " + str(e))

                self.output_file = {
                    'name': output_zip_name + '.zip',
                    'content': zip_buffer.getvalue()
                }

    def prepare_robot_file(rows_data) -> Union[str, None]:
        """
        This function takes the content of the Normalization planning template as input to create
        a csv file that contains the required configuration for the robot execution of the
        normalization in the lab.

        Args:
            row_data: A list of row_data extracted by the importer and already validated by the row_handler.
                      The row_data content will be modified and the robot barcode and coordinates will
                      be added for input and output.
        Returns:
            A string containing the path to the created robot file or None if an error occured.
        """
        def get_destination_container_barcode(row_data):
            return row_data["Destination Container Barcode"]
        def get_destination_contairer_coord(row_data):
            return row_data["Destination Container Barcode"]
        
        def get_robot_destination_container(row_data):
            return row_data["Robot Destination Container"]
        def get_robot_destination_coord(row_data):
            return row_data["Robot Destination Coord"]

        def convert_to_numerical_robot_coord(coord_spec, fms_coord) -> int:
            FIRST_COORD_AXIS = 0
            first_axis_len = len(coord_spec[FIRST_COORD_AXIS])
            second_axis_coord = int(fms_coord[1:])
            return (ord(fms_coord[:1]) - 96) + (first_axis_len * (second_axis_coord - 1))

        mapping_dest_containers = {}
        mapping_src_containers = {}
        coord_spec_by_barcode = {}

        dest_containers = set(row_data["Destination Container Barcode"] for row_data in rows_data)
        dest_coords = list(row_data["Destination Container Coord"] for row_data in rows_data)

        # Map container spec to destination container barcode
        for barcode in dest_containers:
            container = Container.objects.get(barcode=barcode)
            coord_spec_by_barcode[barcode] = CONTAINER_KIND_SPECS[container.kind].coordinate_spec

        # Map destination container barcode to robot destination barcodes
        for i, barcode in enumerate(dest_containers):
            mapping_dest_containers[barcode] = "dest" + str(i)

        # Add robot barcode to the rows_data
        for row_data in rows_data:
            row_data["Robot Destination Container"] = mapping_dest_containers[row_data["Destination Container Barcode"]]

        # Check if robot destinations are plates or tube. Plates have set positions while tubes can be set to maximize 
        # the efficiency of the robot by moving around the input.
        if all(dest_coords):
            # condition plates
            # Add robot dest coord to the rows_data
            for row_data in rows_data:
                row_data["Robot Destination Coord"] = convert_to_numerical_robot_coord(coord_spec_by_barcode(row_data["Destination Container Barcode"]),
                                                                                       row_data["Destination Container Coord"])
        
            sorted_by_robot_container_and_coord = sorted(rows_data,
                                                         key=lambda x: (get_robot_destination_container(x), get_robot_destination_coord(x)),
                                                         reverse=False)

            src_containers = set(row_data["Source Container Barcode"] for row_data in sorted_by_robot_container_and_coord)
            src_coords = list(row_data["Source Container Coord"] for row_data in sorted_by_robot_container_and_coord)

            # Map container spec to source container barcode
            for barcode in src_containers:
                container = Container.objects.get(barcode=barcode)
                coord_spec_by_barcode[barcode] = CONTAINER_KIND_SPECS[container.kind].coordinate_spec

            if all(src_coords):
                # Map destination container barcode to robot destination barcodes
                for i, barcode in enumerate(src_containers):
                    mapping_src_containers[barcode] = "src" + str(i)

                # Add robot src coord to the sorted rows_data
                for row_data in sorted_by_robot_container_and_coord:
                    row_data["Robot Source Coord"] = convert_to_numerical_robot_coord(coord_spec_by_barcode(row_data["Source Container Barcode"]),
                                                                                      row_data["Source Container Coord"])
