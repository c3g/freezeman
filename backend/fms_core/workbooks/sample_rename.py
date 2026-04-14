from typing import Sequence
from ._generic import SheetInfo, TemplateWorkbook
from openpyxl.styles import Color, PatternFill, Font, Alignment
from openpyxl.cell.cell import Cell

from fms_core.services.workbook_utils import CD
from fms_core.templates import SampleRenameHeaders

SHEET_NAMES = ["SampleRename"]

class SampleRenameWorkbook(TemplateWorkbook):
    def __init__(self, sheets_info: Sequence[SheetInfo]):
        super().__init__(sheets_info)

        def style_title(cell: Cell):
            cell.font = Font(name="Calibri", sz=15, family=2, b=True, i=False, color=Color(theme=1), scheme="minor")

        HEADER_PATTERN_FILL = PatternFill(start_color="92d050", end_color="92d050", fill_type="solid")
        def style_section_name(cell: Cell):
            cell.fill = HEADER_PATTERN_FILL

        def red_text(cell: Cell):
            cell.font = Font(name="Calibri", sz=11, family=2, b=True, i=False, color=Color("ff0000"), scheme="minor")
        red_asterisk = CD(value='*', apply_cell=red_text)

        empty_cell = CD(value='')

        self.insert_cells(
            first_cell_location=(1, 1),
            order="row",
            descriptors=[
                [
                    CD(value="Sample Rename Template", apply_cell=style_title)
                ],
                [
                    CD(value="Version : 5.6.0"),
                ],
                [],
                [
                    CD("- Only use the following characters for Sample Name and Sample Alias: a-z, A-Z, 0-9, underscore (_), hyphen (-)")
                ],
                [
                    CD("- Run Processing uses the alias of a sample to identify the sample, not the name.")
                ],
                [],
                [
                    red_asterisk, # barcode
                    empty_cell, # coordinates
                    empty_cell, # index name
                    empty_cell, # old name
                    empty_cell, # old alias
                    red_asterisk, # new name
                    red_asterisk, # new alias
                ],
                [
                    CD(
                        value=SampleRenameHeaders.CONTAINER_BARCODE,
                        apply_cell=style_section_name,
                    ),
                    CD(
                        value=SampleRenameHeaders.CONTAINER_COORDINATES,
                        apply_cell=style_section_name,
                        comment="To identify a sample in a rack, plate, box, etc.",
                    ),
                    CD(
                        value=SampleRenameHeaders.INDEX_NAME,
                        apply_cell=style_section_name,
                        comment="To identify a sample within a pool",
                    ),
                    CD(
                        value=SampleRenameHeaders.OLD_SAMPLE_NAME,
                        apply_cell=style_section_name,
                    ),
                    CD(
                        value=SampleRenameHeaders.OLD_SAMPLE_ALIAS,
                        apply_cell=style_section_name,
                    ),
                    CD(
                        value=SampleRenameHeaders.NEW_SAMPLE_NAME,
                        apply_cell=style_section_name,
                        comment="Changing the name will only affect Freezeman and not the files generated during run processing.",
                    ),
                    CD(
                        value=SampleRenameHeaders.NEW_SAMPLE_ALIAS,
                        apply_cell=style_section_name,
                        comment="Changing the alias will affect the files generated during run processing.",
                    ),
                ],
            ],
            sheet_name=SHEET_NAMES[0],
        )

        center_alignment = Alignment(horizontal="center")

        ws = self.get_sheet_by_name(SHEET_NAMES[0])
        ws['A7'].alignment = center_alignment # barcode
        ws['F7'].alignment = center_alignment # new name
        ws['G7'].alignment = center_alignment # new alias

        for header_name in self.sheets_info[0]['headers']:
            self.set_column_width(header=header_name, width_cm=6.60, sheet_name=SHEET_NAMES[0])

    def headers_row_number(self, sheet_name: str | None = None) -> int:
        return 8
