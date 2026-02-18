from collections.abc import Iterable
from typing import Generic, Literal, Sequence, TypeVar, TypedDict
from openpyxl import Workbook

from fms_core.services.workbook_utils import CellDescription, insert_cells

class SheetInfo(TypedDict):
    name: str
    headers: Sequence[str]

class TemplateWorkbook(Workbook):
    def __init__(self, sheets_info: Sequence[SheetInfo]):
        super().__init__()
        self.sheets_info = sheets_info
        for sheet_info in sheets_info:
            self.create_sheet(title=sheet_info["name"])
        del self["Sheet"]
        self.active = 0

    # beware that column numbers are 1-indexed when using openpyxl, but 0-indexed in the HEADERS list
    def header_to_column_number(self, header: str, sheet_name: str | None = None) -> int:
        """
        Given a header name, returns the corresponding column number in the worksheet (1-indexed).
        
        :returns: the corresponding column number (1-indexed) in the worksheet
        """
        return self.headers(sheet_name=sheet_name).index(header) + 1

    def set_column_width(self, header: str, width_cm: float, sheet_name: str | None = None) -> None:
        col_num = self.header_to_column_number(header, sheet_name=sheet_name)
        col_ord = chr(ord('A') + col_num - 1)
        CM_TO_WHATEVER = 4.489795918
        sheet = self.get_sheet_helper(sheet_name)
        sheet.column_dimensions[col_ord].width = width_cm * CM_TO_WHATEVER

    def headers_row_number(self, sheet_name: str | None = None) -> int:
        """
        :return: 1-indexed row number where headers are located
        """
        raise NotImplementedError("Subclasses of TemplateWorkbook must implement headers_row_number() to specify where headers are located in the worksheet")

    def insert_cells(self, first_cell_location: tuple[int, int], descriptors: Iterable[Iterable[CellDescription]], order: Literal["row", "col"], sheet_name: str | None = None):
        insert_cells(
            worksheet=self.get_sheet_helper(sheet_name),
            first_cell_location=first_cell_location,
            descriptors=descriptors,
            order=order,
        )

    def set_row(self, row_num: int, row_data: dict[str, str], sheet_name: str | None = None):
        sheet = self.get_sheet_helper(sheet_name)
        for col_num, header_name in enumerate(self.headers(), start=1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = row_data[header_name]
    
    def set_rows(self, start_row_num: int, rows_data: Sequence[dict[str, str]], sheet_name: str | None = None):
        for i, row_data in enumerate(rows_data):
            self.set_row(row_num=start_row_num + i, row_data=row_data, sheet_name=sheet_name)
    
    def get_sheet_helper(self, sheet_name: str | None = None):
        return self[sheet_name] if sheet_name else self.active

    def headers(self, sheet_name: str | None = None) -> Sequence[str]:
        if sheet_name is None:
            sheet_name = self.active.title
        for sheet_info in self.sheets_info:
            if sheet_info["name"] == sheet_name:
                return sheet_info["headers"]
        raise ValueError(f"Sheet name {sheet_name} not found in sheets_info")
