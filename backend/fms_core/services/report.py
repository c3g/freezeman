from typing import List
import datetime
from collections import defaultdict

from typing import TypedDict, Union

import pandas as pd
from django.apps import apps
from django.db.models import F, Count,  Sum, Max, Min, TextChoices, functions, QuerySet

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from fms_report.models import Report, MetricField
from fms_report.models._constants import AggregationType, FieldDataType

class TimeWindow(TextChoices):
    MONTHLY = "month", "Monthly"
    WEEKLY = "week", "Weekly"
    DAILY = "day", "Daily"

REPORT_HEADER_COLOR = "34a8eb"

class EnhancedName(TypedDict):
    name: str
    display_name: str

class ReportInfo(TypedDict):
    report: EnhancedName
    groups: list[EnhancedName]
    time_windows: list[str]

class ReportHeader(TypedDict):
    name: str
    display_name: str
    field_order: int
    aggregation: str
    data_type: str

class TimeWindowData(TypedDict):
    time_window: datetime.date
    time_window_label: str
    time_window_start: datetime.date
    time_window_end: datetime.date
    time_window_data: list[dict[str, Union[str, int]]]
    
class ReportData(TypedDict):
    report: EnhancedName
    start_date: str
    end_date: str
    time_window: str
    grouped_by: list[str]
    headers: list[ReportHeader]
    data: list[TimeWindowData]
    

def get_date_range_with_window(start_date: str, end_date: str, time_window: TimeWindow) -> tuple[list[str], list[str]]:
    """
    Provides two lists of dates. One list with each date between two given dates and one list with the matching time window date identifier for each of these dates.

    Args:
        `start_date`: start date for the list of dates.
        `end_date`: end date for the list of dates.
        `time_window`: size of the report page in time.
  
    Returns:
        Tuple with one ordered date list and one matching time window list.
    """
    time_series = pd.date_range(start=start_date, end=end_date).to_series()
    match time_window:
        case TimeWindow.MONTHLY: 
            window_time_series = time_series - pd.to_timedelta(time_series.dt.day - 1, unit="D")
        case TimeWindow.WEEKLY:
            window_time_series = time_series - pd.to_timedelta(time_series.dt.dayofweek, unit="D")
        case TimeWindow.DAILY:
            window_time_series = time_series

    date_range = [datetime.date().isoformat() for datetime in pd.to_datetime(time_series, unit='D')]
    time_window_range = [datetime.date().isoformat() for datetime in pd.to_datetime(window_time_series, unit='D')]
    return date_range, time_window_range


def human_readable_time_window(date: str, time_window: TimeWindow) -> str:
    """
    Provides a human readable label to describe a time window.

    Args:
        `date`: date for which to provide a label.
        `time_window`: size of the report page in time.
  
    Returns:
        String label that best describes the time window to which the date belong.
    """
    time_window_label = None
    date_obj = datetime.date.fromisoformat(date)
    match time_window:
        case TimeWindow.MONTHLY:
            time_window_label = f"{date_obj.strftime('%B')} {date_obj.year}"
        case TimeWindow.WEEKLY:
            time_window_label = f"Week-{date_obj.isocalendar()[1]:02} {date_obj.isocalendar()[0]}"
        case TimeWindow.DAILY:
            time_window_label = date_obj.isoformat()
    return time_window_label


def get_report_as_excel(report_data) -> bytes:
    """
    Converts the report data to an excel workbook bytes stream to be returned to the user.

    Args:
        `report_data`: ReportData for the requested report.

    Returns:
        Bytes stream of an excel workbook.
    """
    out_stream = BytesIO()
    workbook = Workbook()

    headers = report_data["headers"]
    ordered_headers = sorted(headers, key= lambda a : a["field_order"])
    for i, datum in enumerate(report_data["data"], start=1):
        samplesheet_name = datum.get("time_window_label", str(i))
        workbook.create_sheet(samplesheet_name)
        samplesheet = workbook[samplesheet_name]

        # fill header
        for j, header in enumerate(ordered_headers, start=1):
            cell = samplesheet.cell(row=1, column=j)
            header_label = header["display_name"]
            # append aggregation to aggregated fields
            if header.get("aggregation", None) is not None:
                header_label = header_label + f" ({header['aggregation']})"
            cell.value = header_label
            cell.fill = PatternFill(start_color=REPORT_HEADER_COLOR, end_color=REPORT_HEADER_COLOR, fill_type="solid")
        if datum["time_window_data"]:
            for k, sheet_datum in enumerate(datum["time_window_data"], start=2):
                for j, header in enumerate(ordered_headers, start=1):
                    cell = samplesheet.cell(row=k, column=j)
                    cell.value = sheet_datum[header["name"]]
        
    # Remove default sheet
    del workbook["Sheet"]

    workbook.save(out_stream)
    return out_stream.getvalue()


def list_reports() -> list[EnhancedName]:
    """
    Lists all reports that are currently available.

    Returns:
        List of name and display name dictionary for each report.
    """
    return [{"name": values["name"], "display_name": values["display_name"]} for values in Report.objects.all().values("name", "display_name")]


def list_report_information(report_name: str) -> ReportInfo:
    """
    List additional information about a given report.

    Args:
        `name`: Name of the report for which additional information is requested.

    Returns:
        Report info dictionary with important information to define requested report.
    """
    queryset = MetricField.objects.filter(report__name=report_name).all()
    groups = [{"name": field.name, "display_name": field.display_name} for field in queryset if field.is_group]
    time_windows = [TimeWindow.DAILY.label, TimeWindow.WEEKLY.label, TimeWindow.MONTHLY.label]
    report_info = {"report": Report.objects.filter(name=report_name).values("name", "display_name").first(),
                   "groups": groups,
                   "time_windows": time_windows}
    return report_info


def get_report(report_name: str, grouped_by: List[str], time_window: TimeWindow, start_date: str, end_date: str) -> ReportData:
    report_data = {}
    headers = []
    queryset = _get_queryset(report_name, start_date, end_date, time_window, grouped_by)
    report_by_time_window = defaultdict(list)
    for entry in queryset:
        current_row = {key: value for key, value in entry.items() if not key=="time_window"}
        report_by_time_window[entry["time_window"]].append(current_row)
  
    # Creating header
    fields = [column for column in queryset.first().keys() if not column=="time_window"] if queryset.first() is not None else []
    if len(grouped_by) == 0:
        headers = [{**field, "aggregation": None} for field in MetricField.objects.filter(name__in=fields).values("name", "display_name", "field_order", "data_type")]
    else:
        field_ordering_dict = {}
        for field in MetricField.objects.filter(name__in=fields).values("name", "display_name", "field_order", "aggregation", "data_type"):
            field_ordering_dict[field["name"]] = field
        # order for grouping fields
        for i, name in enumerate(grouped_by, start=1):
            field_ordering_dict[name]["field_order"] = i
            headers.append(field_ordering_dict[name])
        # order for value fields
        aggregate_fields = MetricField.objects.filter(name__in=fields).filter(aggregation__isnull=False).values("name", "display_name", "field_order", "aggregation").order_by("field_order")
        for i, field in enumerate(aggregate_fields, start=len(grouped_by)+1):
            field_ordering_dict[field["name"]]["field_order"] = i
            field_ordering_dict[field["name"]]["data_type"] = FieldDataType.NUMBER # Convert data types to numbers for aggregated columns
            headers.append(field_ordering_dict[field["name"]])

    report_data = {
        "report": Report.objects.filter(name=report_name).values("name", "display_name").first(),
        "start_date": start_date,
        "end_date": end_date,
        "time_window": time_window.label,
        "grouped_by": grouped_by,
        "headers": headers,
    }

    # spliting data by time_window
    date_range, date_time_windows = get_date_range_with_window(start_date, end_date, time_window)
    data = []
    current_data = {}
    for date, window in zip(date_range, date_time_windows):
        current_window = current_data.get("time_window", None)
        if current_window is not None and not current_window == window:
                data.append(current_data)
                current_data = {}
        current_data["time_window"] = window
        if current_data.get("time_window_start", None) is None:
            current_data["time_window_label"] = human_readable_time_window(date=window, time_window=time_window)
            current_data["time_window_start"] = date
            current_data["time_window_data"] = report_by_time_window.get(datetime.date.fromisoformat(window), [])
        current_data["time_window_end"] = date
    data.append(current_data)
    report_data["data"] = data

    return report_data


def _get_queryset(report_name: str, start_date: str, end_date: str, time_window: TimeWindow, grouped_by: List[str]) -> QuerySet:
    """
    Provides for each report the basic report quesyset

    Args:
        `name`: name of the report
        `start_date`: start of the report time period requested
        `end_date`: end of the report time period requested
        `time_range`: size of the report page in time.
        `grouped_by`: fields that drive the aggregation of data in order.
    
    Returns:
        Queryset for the report data requested.
    """
    
    # For now this assumes only one field is designated as is_date by report. Once we set more date fields we would need to have the date field name as input.
    report_data = MetricField.objects.filter(report__name=report_name).annotate(date_field=F("name")).filter(is_date=True).values("report__data_model", "date_field")[:1]
    if report_data:
        DataModel = apps.get_model("fms_report", report_data[0]["report__data_model"])

        queryset = ( DataModel.objects.annotate(date_field=F(report_data[0]["date_field"]))
                                      .annotate(time_window=functions.Trunc(F("date_field"), time_window.value))
                                      .filter(date_field__gte=start_date)
                                      .filter(date_field__lte=end_date)
                                      .distinct() )
        
        if (len(grouped_by) == 0):
            # detailed fields
            custom_fields = ["run_name", "lane", "project"]
            ordering_fields = ["time_window", "date_field"]
            ordering_fields.extend(custom_fields)
            detailed_fields = ["time_window"]
            detailed_fields.extend(MetricField.objects.filter(report__name=report_name).values_list("name", flat=True))
            queryset = queryset.values(*detailed_fields)
            queryset = queryset.order_by(*ordering_fields)
        else:
            # grouping definition
            extended_grouped_by = ["time_window"]
            extended_grouped_by.extend(grouped_by)
            queryset = queryset.values(*extended_grouped_by)
            # aggregated fields
            aggregate_fields = MetricField.objects.filter(report__name=report_name).filter(aggregation__isnull=False).values_list("name", "aggregation")
            for name, aggregation in aggregate_fields:
                aggregate = None
                match aggregation:
                    case AggregationType.COUNT:
                        aggregate = Count(F(name), distinct=True)
                    case AggregationType.SUM:
                        aggregate = Sum(F(name))
                    case AggregationType.MAX:
                        aggregate = Max(F(name))
                    case AggregationType.MIN:
                        aggregate = Min(F(name))
                annotation = { f"{name}": aggregate}
                queryset = queryset.annotate(**annotation)
                queryset = queryset.order_by(*extended_grouped_by)
    else:
        queryset = None

    return queryset