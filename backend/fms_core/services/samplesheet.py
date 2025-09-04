from dataclasses import dataclass
from django.templatetags.static import static
from django.conf import settings

from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from fms_core.coordinates import convert_alpha_digit_coord_to_ordinal
from fms_core.containers import CONTAINER_KIND_SPECS
from fms_core.models import DerivedBySample, Index
from fms_core.services.workbook_utils import CD, insert_cells
from fms_core.utils import fit_string_with_ellipsis_in_middle

def get_samplesheet(container_kind, placement):
    """
    Prefill a samplesheet V2 using samples index information and return it as a Byte stream.

    Args:
        `container_kind`: The reported container kind sent with the request. Used to calculate lanes.
        `placement`: List holding each lane mapping between the coordinate and the pooled sample occupying the coordinate.
                    exemple: [{"coordinates": "A01", "sample": "1041243"}, {"coordinates": "B01", "sample": "1414432"}, ...]
    
    Returns:
        `samplesheet`: Samplesheet file containing the prefiled samples and index in excel format.
        `errors`     : Errors raised during creation of the samplesheet file.
        `warnings`   : Warnings raised during creation of the samplesheet file.

    """
    out_stream = BytesIO()
    errors = []
    warnings = []

    row_data_by_lane = []
    try:
        container_spec = CONTAINER_KIND_SPECS[container_kind]
        if container_spec.is_run_container:
            if container_spec is None:
                raise Exception(f'Cannot convert coord to lane number. No ContainerSpec found for container kind "{container_kind}".')
            for lane_info in placement:
                coordinates = lane_info["coordinates"]
                sample_id = lane_info["sample_id"]
                lane = convert_alpha_digit_coord_to_ordinal(coordinates, container_spec.coordinate_spec)
                derived_by_samples = DerivedBySample.objects.filter(sample_id=sample_id)
                values = derived_by_samples.values_list("derived_sample__biosample__alias", "derived_sample__id", "derived_sample__library__index_id")
                for sample_alias, derived_sample_id, index_id in values:
                    if index_id is not None:
                        index = Index.objects.get(id=index_id)
                        sequences_3prime = ", ".join(index.list_3prime_sequences)
                        sequences_5prime = ", ".join(index.list_5prime_sequences)
                        row_data_by_lane.append((str(lane), sample_alias, derived_sample_id, sequences_3prime, sequences_5prime))
                    else:
                        errors.append(f'Cannot find index associated to sample {sample_alias}.')
        else:
            errors.append(f"Container of kind {container_kind} cannot be used for an experiment run.")
    except Exception as err:
        errors.append(err)

    row_data_by_lane.sort(key=lambda x: x[0])
    bclconvert_data = [
        BCLConvert_Datum(Lane=lane, Sample_ID=_get_Sample_ID(sample_alias, derived_sample_id), Index=sequences_3prime, Index2=sequences_5prime)
        for lane, sample_alias, derived_sample_id, sequences_3prime, sequences_5prime
        in row_data_by_lane
    ]
    row_data_by_lane.sort(key=lambda x: x[1])
    dragen_data = [
        DragenGermline_Datum(Sample_ID=_get_Sample_ID(sample_alias, derived_sample_id))
        for lane, sample_alias, derived_sample_id, sequences_3prime, sequences_5prime
        in row_data_by_lane
    ]
    workbook = _generate_samplesheet_workbook(
        BCLConvert_Data=bclconvert_data,
        DragenGermline_Data=dragen_data
    )
    workbook.save(out_stream)
    return out_stream.getvalue(), errors, warnings

def _get_Sample_ID(sample_alias, derived_sample_id):
    MAX_SAMPLE_ID_LENGTH = 100
    SEPARATOR = "_"
    sample_alias = str(sample_alias)
    derived_sample_id = str(derived_sample_id)
    max_sample_alias_length = MAX_SAMPLE_ID_LENGTH - len(SEPARATOR) - len(derived_sample_id)

    sample_alias = fit_string_with_ellipsis_in_middle(sample_alias, max_sample_alias_length, "---")

    return f"{sample_alias}_{derived_sample_id}"

@dataclass
class BCLConvert_Datum:
    Lane: str
    Sample_ID: str
    Index: str
    Index2: str

@dataclass
class DragenGermline_Datum:
    Sample_ID: str

def _generate_samplesheet_workbook(BCLConvert_Data: list[BCLConvert_Datum], DragenGermline_Data: list[DragenGermline_Datum]) -> Workbook:
    workbook = Workbook()

    workbook.create_sheet("Samplesheet")
    samplesheet = workbook["Samplesheet"]
    del workbook["Sheet"]

    workbook.create_sheet("Info")
    workbook.create_sheet("Index")

    # fill patterns
    fillSectionName = PatternFill(start_color="b3cac7", end_color="b3cac7", fill_type="solid")
    def style_section_name(cell):
        cell.fill = fillSectionName
    fillKey = PatternFill(start_color="e8a202", end_color="e8a202", fill_type="solid")
    def style_header(cell):
        cell.fill = fillKey
    fillPrefilled = PatternFill(start_color="dee7e5", end_color="dee7e5", fill_type="solid")
    def style_prefilled(cell):
        cell.fill = fillPrefilled
    fillWritable = PatternFill(start_color="e8f2a1", end_color="e8f2a1", fill_type="solid")
    def style_writable(cell):
        cell.fill = fillWritable
    fillHardcoded = PatternFill(start_color="fc8b8b", end_color="fc8b8b", fill_type="solid")
    def style_hardcoded(cell):
        cell.fill = fillHardcoded

    MAX_COLUMN = 5
    def add_section_header(section_name: str):
        samplesheet.append([f"[{section_name}]"])
        for i in range(1, MAX_COLUMN):
            samplesheet.cell(row=samplesheet.max_row, column=i).fill = fillSectionName

    add_section_header("Header")
    MAX_RUN_NAME_LENGTH = 255
    INSTRUMENT_PLATFORMS = ["NovaSeqXSeries"]
    INDEX_ORIENTATIONS = ["Forward"]
    insert_cells(
        samplesheet,
        first_cell_location=(samplesheet.max_row+1, 1),
        order="col",
        descriptors=[
            [
                CD(value="FileFormatVersion", apply_cell=style_header),
                CD(value="RunName", apply_cell=style_header),
                CD(value="InstrumentPlatform", apply_cell=style_header),
                CD(value="IndexOrientation", apply_cell=style_header)
            ],
            [
                CD(
                    value="2",
                    validation=DataValidation(type="whole", operator="equal", formula1=2, allow_blank=False, showErrorMessage=True, errorTitle="Invalid FileFormatVersion Value", error="FileFormatVersion must always be 2"),
                    apply_cell=style_writable
                ),
                CD(
                    value="",
                    validation=DataValidation(type="textLength", operator="lessThanOrEqual", formula1=MAX_RUN_NAME_LENGTH, allow_blank=True, showErrorMessage=True, errorTitle="Invalid RunName Length", error=f"RunName must be less than or equal to {MAX_RUN_NAME_LENGTH} characters"),
                    apply_cell=style_writable
                ),
                CD(
                    value=INSTRUMENT_PLATFORMS[0],
                    validation=DataValidation(type="list", formula1=f'"{",".join(INSTRUMENT_PLATFORMS)}"', allow_blank=False, showErrorMessage=True, errorTitle="Invalid InstrumentPlatform Value", error=f"Only {', '.join(INSTRUMENT_PLATFORMS)} is supported"),
                    apply_cell=style_writable
                ),
                CD(
                    value=INDEX_ORIENTATIONS[0],
                    validation=DataValidation(type="list", formula1=f'"{",".join(INDEX_ORIENTATIONS)}"', allow_blank=False, showErrorMessage=True, errorTitle="Invalid IndexOrientation Value", error=f"Only {', '.join(INDEX_ORIENTATIONS)} is supported"),
                    apply_cell=style_writable
                )
            ]
        ]
    )

    samplesheet.append([])
    add_section_header("Reads")
    insert_cells(
        samplesheet,
        first_cell_location=(samplesheet.max_row+1, 1),
        order="col",
        descriptors=[
            [
                CD(value="Read1Cycles", apply_cell=style_header),
                CD(value="Read2Cycles", apply_cell=style_header),
                CD(value="Index1Cycles", apply_cell=style_header),
                CD(value="Index2Cycles", apply_cell=style_header)
            ],
            [
                CD(
                    value="151",
                    validation=DataValidation(type="whole", operator="greaterThan", formula1=0, allow_blank=False, showErrorMessage=True, errorTitle="Invalid Read1Cycles Value", error="Read1Cycles must be greater than 0"),
                    apply_cell=style_hardcoded
                ),
                CD(
                    value="151",
                    validation=DataValidation(type="whole", operator="greaterThan", formula1=0, allow_blank=False, showErrorMessage=True, errorTitle="Invalid Read2Cycles Value", error="Read2Cycles must be greater than 0"),
                    apply_cell=style_hardcoded
                ),
                CD(
                    value="10",
                    validation=DataValidation(type="whole", operator="greaterThanOrEqual", formula1=0, allow_blank=False, showErrorMessage=True, errorTitle="Invalid Index1Cycles Value", error="Index1Cycles must be greater than or equal to 0"),
                    apply_cell=style_hardcoded
                ),
                CD(
                    value="10",
                    validation=DataValidation(type="whole", operator="greaterThanOrEqual", formula1=0, allow_blank=False, showErrorMessage=True, errorTitle="Invalid Index2Cycles Value", error="Index2Cycles must be greater than or equal to 0"),
                    apply_cell=style_hardcoded
                ),
            ]
        ]
    )

    samplesheet.append([])
    add_section_header("Sequencing_Settings")
    LIBRARY_PREP_KITS = ["IlluminaDNAPCRFree"]
    insert_cells(
        samplesheet,
        first_cell_location=(samplesheet.max_row+1, 1),
        order="col",
        descriptors=[
            [CD(value="LibraryPrepKits", apply_cell=style_header, comment="Leave it blank unless you want to fill the [Cloud_Settings] section.")],
            [
                CD(
                    value="",
                    validation=DataValidation(type="list", formula1=f'"{",".join(LIBRARY_PREP_KITS)}"', allow_blank=True, showErrorMessage=True, errorTitle="Invalid LibraryPrepKits Value", error=f"Only {', '.join(LIBRARY_PREP_KITS)} is supported"),
                    apply_cell=style_writable
                )
            ]
        ]
    )

    samplesheet.append([])
    add_section_header("BCLConvert_Settings")
    ADAPTERREAD1S = { "IlluminaDNAPCRFree": "CTGTCTCTTATACACATCT+ATGTGTATAAGAGACA" }
    ADAPTERREAD2S = { "IlluminaDNAPCRFree": "CTGTCTCTTATACACATCT+ATGTGTATAAGAGACA" }
    FASTQ_COMPRESSION_FORMATS = ["gzip"]
    insert_cells(
        samplesheet,
        first_cell_location=(samplesheet.max_row+1, 1),
        order="col",
        descriptors=[
            [
                CD(value="SoftwareVersion", apply_cell=style_header),
                CD(value="AdapterRead1", apply_cell=style_header),
                CD(value="AdapterRead2", apply_cell=style_header),
                CD(value="OverrideCycles", apply_cell=style_header),
                CD(value="FastqCompressionFormat", apply_cell=style_header)
            ],
            [
                CD(value="4.3.13", apply_cell=style_hardcoded),
                CD(value=ADAPTERREAD1S["IlluminaDNAPCRFree"], apply_cell=style_writable),
                CD(value=ADAPTERREAD2S["IlluminaDNAPCRFree"], apply_cell=style_writable),
                CD(value="Y151;I10;I10;Y151", apply_cell=style_hardcoded),
                CD(
                    value="gzip",
                    validation=DataValidation(type="list", formula1=f'"{",".join(FASTQ_COMPRESSION_FORMATS)}"', allow_blank=False, showErrorMessage=True, errorTitle="Invalid FastqCompressionFormat Value", error="Only gzip is supported"),
                    apply_cell=style_writable
                )
            ]
        ]
    )

    samplesheet.append([])
    add_section_header("BCLConvert_Data")
    insert_cells(
        samplesheet,
        first_cell_location=(samplesheet.max_row+1, 1),
        order="row",
        descriptors=[
            [
                CD(value="Lane", apply_cell=style_header),
                CD(value="Sample_ID", apply_cell=style_header),
                CD(value="Index", apply_cell=style_header),
                CD(value="Index2", apply_cell=style_header)
            ],
        ] + [
            [
                CD(value=datum.Lane, apply_cell=style_prefilled),
                CD(value=datum.Sample_ID, apply_cell=style_prefilled),
                CD(value=datum.Index, apply_cell=style_prefilled),
                CD(value=datum.Index2, apply_cell=style_prefilled)
            ]
            for datum in BCLConvert_Data
        ]
    )

    samplesheet.append([])
    add_section_header("DragenGermline_Settings")
    MAP_ALIGN_OUT_FORMATS = ["bam", "cram", "none"]
    KEEP_FASTQS = ["TRUE", "FALSE"]
    insert_cells(
        samplesheet,
        first_cell_location=(samplesheet.max_row+1, 1),
        order="col",
        descriptors=[
            [
                CD(value="SoftwareVersion", apply_cell=style_header),
                CD(value="AppVersion", apply_cell=style_header),
                CD(value="MapAlignOutFormat", apply_cell=style_header),
                CD(value="KeepFastq", apply_cell=style_header)
            ],
            [
                CD(value="4.3.13", apply_cell=style_hardcoded),
                CD(value="1.3.11", apply_cell=style_hardcoded),
                CD(
                    value="none",
                    validation=DataValidation(type="list", formula1=f'"{",".join(MAP_ALIGN_OUT_FORMATS)}"', allow_blank=False, showErrorMessage=True, errorTitle="Invalid MapAlignOutFormat Value", error=f"Only {', '.join(MAP_ALIGN_OUT_FORMATS)} are supported"),
                    apply_cell=style_writable
                ),
                CD(
                    value="TRUE",
                    validation=DataValidation(type="list", formula1=f'"{",".join(KEEP_FASTQS)}"', allow_blank=False, showErrorMessage=True, errorTitle="Invalid KeepFastq Value", error=f"Only {', '.join(KEEP_FASTQS)} are supported"),
                    apply_cell=style_writable
                )
            ]
        ]
    )

    samplesheet.append([])
    add_section_header("DragenGermline_Data")
    REFERENCE_GENOME_REFERENCE_DIRS = [
        "hg38-alt_masked.cnv.graph.hla.rna-10-r4.0-2",
        "hg19-alt_masked.cnv.graph.hla.rna-10-r4.0-1",
        "chm13_v2-cnv.graph.hla.rna-10-r4.0-1",
    ]
    REFERENCE_GENOME_VALIDATION = DataValidation(
        type="list",
        formula1=f'"{",".join(REFERENCE_GENOME_REFERENCE_DIRS)}"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid ReferenceGenomeDir Value",
        error=f"Only {', '.join(REFERENCE_GENOME_REFERENCE_DIRS)} are supported"
    )
    VARIANT_CALLING_MODE_VALIDATION = DataValidation(
        type="list",
        formula1='"None,SmallVariantCaller,AllVariantCallers"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid VariantCallingMode Value",
        error="Only None, SmallVariantCaller, AllVariantCallers are supported"
    )
    insert_cells(
        samplesheet,
        first_cell_location=(samplesheet.max_row+1, 1),
        order="row",
        descriptors=[
            [
                CD(value="Sample_ID", apply_cell=style_header),
                CD(value="ReferenceGenomeDir", apply_cell=style_header),
                CD(value="VariantCallingMode", apply_cell=style_header)
            ],
        ] + [
            [
                CD(value=datum.Sample_ID, apply_cell=style_prefilled),
                CD(value=REFERENCE_GENOME_REFERENCE_DIRS[0], validation=REFERENCE_GENOME_VALIDATION, apply_cell=style_writable),
                CD(value="None", validation=VARIANT_CALLING_MODE_VALIDATION, apply_cell=style_writable),
            ]
            for datum in DragenGermline_Data
        ]
    )

    samplesheet.column_dimensions["A"].width = 25
    samplesheet.column_dimensions["B"].width = 20
    samplesheet.column_dimensions["C"].width = 20
    samplesheet.column_dimensions["D"].width = 20
    samplesheet.column_dimensions["E"].width = 20

    # add border to all cells in the bounding range
    bounding_range = samplesheet.calculate_dimension().split(":")
    bounding_range = coordinate_from_string(bounding_range[0]), coordinate_from_string(bounding_range[1])
    bounding_range = (column_index_from_string(bounding_range[0][0]), int(bounding_range[0][1])), (column_index_from_string(bounding_range[1][0]), int(bounding_range[1][1]))
    for i in range(bounding_range[0][0], bounding_range[1][0]+1):
        for j in range(bounding_range[0][1], bounding_range[1][1]+1):
            cell = samplesheet.cell(row=j, column=i)
            cell.border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"),
            )

    infosheet = workbook["Info"]
    # the standard dict type maintains insertion order by default
    INFO = {
        "Header" : {
            "FileFormatVersion": "Used to identify the sample sheet as a v2 sample sheet. Must always be 2.",
            "RunName": "The run name can contain 255 alphanumeric characters, spaces, dashes, and underscores.",
            "InstrumentPlatform": "Identifies the instrument platform to be used for the run. Only NovaSeqXSeries is supported.",
            "IndexOrientation": "We will only support Forward.",
        },
        "Reads": {
            "Read1Cycles": "Number of cycles for Read 1.",
            "Read2Cycles": "Number of cycles for Read 1. Read 2 required if running a paired-end sequencing run.",
            "Index1Cycles": "Number of cycles in Index Read 1. Index 1 required if more than one sample is present in sample sheet.",
            "Index2Cycles": "Number of cycles in Index Read 2. Index 2 required if using dual indexes for demultiplexing.",
        },
        "Sequencing_Settings": {
            "LibraryPrepKits": "Identifies the library prep kit used for the run. Only IlluminaDNAPCRFree is supported.",
        },
        "BCLConvert_Settings": {
            "SoftwareVersion": "Identifies the version of the DRAGEN software to be used process the specific DRAGEN pipeline, including conversion to FASTQ. Use all three integers included in the DRAGEN version name. For example, 3.5.7",
            "AdapterRead1": "The sequence tied to the Library Kits used to prepare the sample. To trim multiple adapters, separate the sequences with a plus sign (+).",
            "AdapterRead2": "The sequence tied to the Library Kits used to prepare the sample. To trim multiple adapters, separate the sequences with a plus sign (+).",
            "OverrideCycles": "Examples of OverrideCycles: U8Y143;I8;I8;U8Y143 or N10Y66;I6;N10Y66",
            "FastqCompressionFormat": "The compression format for the FASTQ output files. Only gzip is supported.",
        },
        "BCLConvert_Data": {
            "Lane": "Specifies FASTQ files only for the samples with the specified lane number.",
            "Sample_ID": "Freezeman formats Sample_ID as [Sample Alias]_[Derived Sample ID]. Sample_IDs can be repeated. This assumes that a sample is split across lanes, and results will be merged for samples with the same ID.",
            "Index": "The Index 1 (i7) index adapter sequence. Index 1 is required if index cycles are specified for the sample in OverrideCycles.",
            "Index2": "The Index 2 (i5) index adapter sequence. Index 2 is required if Index2Cycles is > 0. Index2 is entered in the sample sheet in forward, non-complemented format for user convenience",
        },
        "DragenGermline_Settings": {
            "SoftwareVersion": "Identifies the version of the DRAGEN software to be used process the specific DRAGEN pipeline, including conversion to FASTQ. Use all three integers included in the DRAGEN version name. For example, 3.5.7",
            "AppVersion": "The version of the workflow-specific application. Use all three integers included in the version name. For example, 3.5.7",
            "MapAlignOutFormat": "Formatting of the alignment output files. (bam, cram, none)",
            "KeepFastq": "Indicates whether FASTQ files are saved (TRUE) or discarded (FALSE).",
        },
        "DragenGermline_Data": {
            "Sample_ID": "Freezeman formats Sample_ID as [Sample Alias]_[Derived Sample ID]. Sample_IDs must be unique. If samples are split across lanes, then there must be a [BCLConvert_Data] section with duplicate samples indicating which ones will be merged.",
            "ReferenceGenomeDir": "Dragen requires its own custom binaries for reference genomes.",
            "VariantCallingMode": "None (No variant calling will be performed), SmallVariantCaller (Only small variant calling will be performed), AllVariantCallers (The following variant callers will be run: Small, Structural, CNV, Repeat Expansions, ROH, CYP2D6)",
        }
    }
    info_descriptors = []
    for section_name, section in INFO.items():
        info_descriptors.append([CD(value=f"[{section_name}]", apply_cell=style_section_name)])
        for header, value in section.items():
            info_descriptors.append([CD(value=header, apply_cell=style_header), CD(value=value)])
        info_descriptors.append([CD(value="")])
    insert_cells(
        infosheet,
        first_cell_location=(1, 1),
        order="row",
        descriptors=info_descriptors
    )
    infosheet.append(["To view the full requirements for each section, please consult the documentation: https://help.connected.illumina.com/run-set-up/overview/parameters"])
    infosheet.column_dimensions["A"].width = 25

    return workbook

def fit_sheet_columns_width_to_content(workbook_sheet: Worksheet, adjust_factor: float=1.1, adjust_offset: float=4):
    # Refit columns width to data
    for column in workbook_sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length * adjust_factor) + adjust_offset # Ad Hoc values to tweak the width
        workbook_sheet.column_dimensions[column[0].column_letter].width = adjusted_width