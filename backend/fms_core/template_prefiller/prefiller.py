import os
from io import BytesIO
from openpyxl.reader.excel import load_workbook
from ._utils import load_position_dict, find_worksheet_header_offset
from django.conf import settings


def PrefillTemplate(template_path, template_info, queryset):
    """
    Function that return a prefilled template byte stream.
    It fetch a designated empty template using template_path.
    It loads data from a queryset according to the template_info prefill info attribute.
    It used the template_info sheet info to locate and position the information in the excel workbook.

    position_dict has the following structure :
    {SHEET_NAME: {header_offset: HEADER_OFFSET, queryset_column_list: [COLUMN_NAME, ...], column_offsets: {COLUMN_NAME: COLUMN_OFFSET, ...}}, ...}
    """
    out_stream = BytesIO()

    workbook = load_workbook(filename=template_path)
    position_dict = load_position_dict(workbook, template_info["sheets info"], template_info["prefill info"])

    for sheet_name, sheet_dict in position_dict.items():
        current_sheet = workbook[sheet_name]
        for i, entry in enumerate(queryset.values(*sheet_dict["queryset_column_list"])):
            for prefill_sheet_name, template_column, queryset_column in template_info["prefill info"]:
                if prefill_sheet_name == sheet_name:
                    current_sheet.cell(row=sheet_dict["header_offset"] + i, column=sheet_dict["column_offsets"][template_column]).value = entry[queryset_column]
            
    workbook.save(out_stream)
    return out_stream.getvalue()


def PrefillTemplateFromDict(template, rows_dicts):
    """
    Function that return a prefilled template byte stream.
    It fetch a designated empty template using template_path.
    For each sheet, It loads data from a dictionary with format { column : value } for each row
    It used the template_info sheet info to locate and position the information in the excel workbook.

    rows_dicts is a list of list of row dictionaries: one list for each sheet to prefill, one list for each row.
    The order of the dicts match the one from the template sheets_info.
    """
    # Get the template information
    filename = "/".join(template["identity"]["file"].split("/")[-2:])
    template_path = os.path.join(settings.STATIC_ROOT, filename)
    out_stream = BytesIO()
    workbook = load_workbook(filename=template_path)
    
    # Populate template
    try:
        for i, sheet_info in enumerate(template["sheets info"]):
            current_sheet = workbook[sheet_info["name"]]
            header_offset = find_worksheet_header_offset(current_sheet, sheet_info['headers'])
            for j, entry in enumerate(rows_dicts[i]):
                for header_index, template_column in enumerate(sheet_info['headers']):
                    current_sheet.cell(row=header_offset + j, column=header_index + 1).value = entry.get(template_column, None)
        workbook.save(out_stream)
        return out_stream.getvalue()
    except Exception as e:
        raise e