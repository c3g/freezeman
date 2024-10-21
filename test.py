from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

def add_horizontal_table(worksheet: Worksheet, headers: list[str], data: list[list[str]], topleft: tuple[int, int]):
    header_cells = list[Cell]()
    data_cells = list[list[Cell]]()

    for col_offset, header in enumerate(headers):
        header_cells.append(worksheet.cell(row=topleft[0], column=topleft[1]+col_offset, value=header))
    for row_offset, row_data in enumerate(data):
        row_cells = list[Cell]()
        for col_offset, value in enumerate(row_data):
            # +1 to skip the header row
            row_cells.append(worksheet.cell(row=topleft[0]+row_offset+1, column=topleft[1]+col_offset, value=value))
        data_cells.append(row_cells)

    return header_cells, data_cells

def add_vertical_table(worksheet: Worksheet, headers: list[str], data: list[list[str]], topleft: tuple[int, int]):
    header_cells = list[Cell]()
    data_cells = list[list[Cell]]()

    for row_rel_index, header in enumerate(headers):
        header_cells.append(worksheet.cell(row=topleft[0]+row_rel_index, column=topleft[1], value=header))
    for col_rel_index, col_data in enumerate(data):
        col_cells = list[Cell]()
        for row_rel_index, value in enumerate(col_data):
            # +1 to skip the header column
            col_cells.append(worksheet.cell(row=topleft[0]+row_rel_index, column=topleft[1]+col_rel_index+1, value=value))
        data_cells.append(col_cells)

    return header_cells, data_cells

@dataclass
class HorizontalTable:
    headers: list[Cell]
    data: list[list[Cell]]

@dataclass
class BCLConvert_Datum:
    Lane: str
    Sample_ID: str
    Index: str
    Index2: str

@dataclass
class DragenGermline_Datum:
    Sample_ID: str
    ReferenceGenomeDir: str
    VariantCallingMode: str

def generate_samplesheet_workbook(BCLConvert_Data: list[BCLConvert_Datum], DragenGermline_Data: list[DragenGermline_Datum]) -> Workbook:
    workbook = Workbook()

    workbook.create_sheet("Samplesheet")
    samplesheet = workbook["Samplesheet"]
    del workbook["Sheet"]

    workbook.create_sheet("Info")
    workbook.create_sheet("Index")

    MAX_COLUMN = 5
    fillSectionName = PatternFill(start_color="b3cac7", end_color="b3cac7", fill_type="solid")
    fillKey = PatternFill(start_color="e8a202", end_color="e8a202", fill_type="solid")
    fillPrefilled = PatternFill(start_color="dee7e5", end_color="dee7e5", fill_type="solid")
    fillBlank = PatternFill(start_color="e8f2a1", end_color="e8f2a1", fill_type="solid")

    def add_section_header(section_name: str):
        samplesheet.append([f"[{section_name}]"])
        for i in range(1, MAX_COLUMN):
            samplesheet.cell(row=samplesheet.max_row, column=i).fill = fillSectionName

    MAX_RUN_NAME_LENGTH = 255
    add_section_header("Header")
    header_cells, _ = add_vertical_table(
        samplesheet,
        headers=["FileFormatVersion", "RunName", "InstrumentPlatform", "IndexOrientation"],
        data=[["@FileFormatVersion", "@RunName", "@InstrumentPlatform", "@IndexOrientation"]],
        topleft=(samplesheet.max_row+1, 1)
    )
    for cell in header_cells:
        cell.fill = fillKey

    samplesheet.append([])
    add_section_header("Reads")
    header_cells, _ = add_vertical_table(
        samplesheet,
        headers=["Read1Cycles", "Read2Cycles", "Index1Cycles", "Index2Cycles"],
        data=[["@Read1Cycles", "@Read2Cycles", "@Index1Cycles", "@Index2Cycles"]],
        topleft=(samplesheet.max_row+1, 1)
    )
    for cell in header_cells:
        cell.fill = fillKey

    samplesheet.append([])
    add_section_header("Sequencing_Settings")
    header_cells, data_cells = add_vertical_table(
        samplesheet,
        headers=["LibraryPrepKits"],
        data=[["@LibraryPrepKits"]],
        topleft=(samplesheet.max_row+1, 1)
    )
    header_cells[0].fill = fillKey

    samplesheet.append([])
    add_section_header("BCLConvert_Settings")
    header_cells, data_cells = add_vertical_table(
        samplesheet,
        headers=["SoftwareVersion", "AdapterRead1", "AdapterRead2", "OverrideCycles", "FastqCompressionFormat"],
        data=[["@BCLConvert_SoftwareVersion", "@AdapterRead1", "@AdapterRead2", "@OverrideCycles", "@FastqCompressionFormat"]],
        topleft=(samplesheet.max_row+1, 1)
    )
    for cell in header_cells:
        cell.fill = fillKey
    adapter_read_cells = [data_cells[0][1], data_cells[0][2]]

    samplesheet.append([])
    add_section_header("BCLConvert_Data")
    header_cells, data_cells = add_horizontal_table(
        samplesheet,
        headers=["Lane", "Sample_ID", "Index", "Index2"],
        data=[[datum.Lane, datum.Sample_ID, datum.Index, datum.Index2] for datum in BCLConvert_Data],
        topleft=(samplesheet.max_row+1, 1)
    )
    for cell in header_cells:
        cell.fill = fillKey
    for row_cells in data_cells:
        for cell in row_cells:
            cell.fill = fillPrefilled

    samplesheet.append([])
    add_section_header("DragenGermline_Settings")
    header_cells, _ = add_vertical_table(
        samplesheet,
        headers=["SoftwareVersion", "AppVersion", "MapAlignOutFormat", "KeepFastq"],
        data=[["@DragenGermline_SoftwareVersion", "@AppVersion", "@MapAlignOutFormat", "@KeepFastq"]],
        topleft=(samplesheet.max_row+1, 1)
    )
    for cell in header_cells:
        cell.fill = fillKey

    samplesheet.append([])
    add_section_header("DragenGermline_Data")
    header_cells, data_cells = add_horizontal_table(
        samplesheet,
        headers=["Sample_ID", "ReferenceGenomeDir", "VariantCallingMode"],
        data=[[datum.Sample_ID, datum.ReferenceGenomeDir, datum.VariantCallingMode] for datum in DragenGermline_Data],
        topleft=(samplesheet.max_row+1, 1)
    )
    for cell in header_cells:
        cell.fill = fillKey
    for row_cells in data_cells:
        row_cells[0].fill = fillPrefilled

    samplesheet_cells_validations: dict[str, list[DataValidation]] = {
        # [Header]
        "FileFormatVersion": [
            DataValidation(type="whole", operator="equal", formula1=2, allow_blank=False, showErrorMessage=True, errorTitle="Invalid FileFormatVersion Value", error="FileFormatVersion must always be 2")
        ],
        "RunName": [
            DataValidation(type="textLength", operator="lessThanOrEqual", formula1=MAX_RUN_NAME_LENGTH, allow_blank=True, showErrorMessage=True, errorTitle="Invalid RunName Length", error=f"RunName must be less than or equal to {MAX_RUN_NAME_LENGTH} characters")
        ],
        "InstrumentPlatform": [
            DataValidation(type="list", formula1='"NovaSeqXSeries"', allow_blank=False, showErrorMessage=True, errorTitle="Invalid InstrumentPlatform Value", error="Only NovaSeqXSeries is supported")
        ],
        # "InstrumentType": [],
        "IndexOrientation": [
            DataValidation(type="list", formula1='"Forward"', allow_blank=False, showErrorMessage=True, errorTitle="Invalid IndexOrientation Value", error="Only Forward is supported")
        ],
        # [Reads]
        "Read1Cycles": [
            DataValidation(type="whole", operator="greaterThan", formula1=0, allow_blank=False, showErrorMessage=True, errorTitle="Invalid Read1Cycles Value", error="Read1Cycles must be greater than 0")
        ],
        "Read2Cycles": [
            DataValidation(type="whole", operator="greaterThan", formula1=0, allow_blank=False, showErrorMessage=True, errorTitle="Invalid Read2Cycles Value", error="Read2Cycles must be greater than 0")
        ],
        "Index1Cycles": [
            DataValidation(type="whole", operator="greaterThanOrEqual", formula1=0, allow_blank=False, showErrorMessage=True, errorTitle="Invalid Index1Cycles Value", error="Index1Cycles must be greater than or equal to 0")
        ],
        "Index2Cycles": [
            DataValidation(type="whole", operator="greaterThanOrEqual", formula1=0, allow_blank=False, showErrorMessage=True, errorTitle="Invalid Index2Cycles Value", error="Index2Cycles must be greater than or equal to 0")
        ],
        # [Sequencing_Settings]
        "LibraryPrepKits": [
            DataValidation(type="list", formula1='"IlluminaDNAPCRFree"', allow_blank=False, showErrorMessage=True, errorTitle="Invalid LibraryPrepKits Value", error="Only IlluminaDNAPCRFree is supported")
        ],
        # [BCLConvert_Settings]
        "BCLConvert_SoftwareVersion": [],
        "AdapterRead1": [],
        "AdapterRead2": [],
        "OverrideCycles": [],
        "FastqCompressionFormat": [
            DataValidation(type="list", formula1='"gzip"', allow_blank=False, showErrorMessage=True, errorTitle="Invalid FastqCompressionFormat Value", error="Only gzip is supported")
        ],
        # [BCLConvert_Data]
        # [DragenGermline_Settings]
        "DragenGermline_SoftwareVersion": [],
        "AppVersion": [],
        "MapAlignOutFormat": [
            DataValidation(type="list", formula1='"bam,cram,none"', allow_blank=False, showErrorMessage=True, errorTitle="Invalid MapAlignOutFormat Value", error="Only bam, cram, none are supported")
        ],
        "KeepFastq": [
            DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False, showErrorMessage=True, errorTitle="Invalid KeepFastq Value", error="Only TRUE, FALSE are supported")
        ],
        # [DragenGermline_Data]
        "ReferenceGenomeDir": [],
        "VariantCallingMode": [
            DataValidation(type="list", formula1='"None,SmallVariantCaller,AllVariantCallers"', allow_blank=False, showErrorMessage=True, errorTitle="Invalid VariantCallingMode Value", error="Only None, SmallVariantCaller, AllVariantCallers are supported")
        ]
    }
    samplesheet_cells_default_values: dict[str, str] = {
        # [Header]
        "FileFormatVersion": "2",
        "RunName": "",
        "InstrumentPlatform": "NovaSeqXSeries",
        "InstrumentType": "",
        "IndexOrientation": "Forward",
        # [Reads]
        "Read1Cycles": "",
        "Read2Cycles": "",
        "Index1Cycles": "",
        "Index2Cycles": "",
        # [Sequencing_Settings]
        "LibraryPrepKits": "IlluminaDNAPCRFree",
        # [BCLConvert_Settings]
        "BCLConvert_SoftwareVersion": "",
        "AdapterRead1": "CTGTCTCTTATACACATCT+ATGTGTATAAGAGACA",
        "AdapterRead2": "CTGTCTCTTATACACATCT+ATGTGTATAAGAGACA",
        "OverrideCycles": "",
        "FastqCompressionFormat": "gzip",
        # [BCLConvert_Data]
        # [DragenGermline_Settings]
        "DragenGermline_SoftwareVersion": "",
        "AppVersion": "",
        "MapAlignOutFormat": "none",
        "KeepFastq": "TRUE",
        # [DragenGermline_Data]
        "ReferenceGenomeDir": "",
        "VariantCallingMode": "None"
    }
    for row_cells in samplesheet.rows:
        for cell in row_cells:
            value = cell.value
            if value and value.startswith("@"):
                key = value[1:]
                cell.value = samplesheet_cells_default_values.get(key, "UNKNOWN_DEFAULT_VALUE")
                cell.fill = fillBlank
                for validation in samplesheet_cells_validations.get(key, []):
                    samplesheet.add_data_validation(validation)
                    validation.add(cell)

    samplesheet.column_dimensions["A"].width = 25
    samplesheet.column_dimensions["B"].width = 20
    samplesheet.column_dimensions["C"].width = 20
    samplesheet.column_dimensions["D"].width = 20
    samplesheet.column_dimensions["E"].width = 20

    # add border to all cells in the bounding range
    bounding_range = samplesheet.calculate_dimension().split(":")
    bounding_range = coordinate_from_string(bounding_range[0]), coordinate_from_string(bounding_range[1])
    bounding_range = (column_index_from_string(bounding_range[0][0]), int(bounding_range[0][1])), (column_index_from_string(bounding_range[1][0]), int(bounding_range[1][1]))
    for i in range(bounding_range[0][0], bounding_range[1][0]+1):
        for j in range(bounding_range[0][1], bounding_range[1][1]+1):
            cell = samplesheet.cell(row=j, column=i)
            cell.border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"),
            )

    # erase border on the right of adapter read value cells
    for adapter_read_cell in adapter_read_cells:
        adapter_read_cell.border = Border(
            left=Side(border_style="thin", color="000000"),
            right=None,
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        right_cell = adapter_read_cell.row, adapter_read_cell.column
        right_cell = right_cell[0], right_cell[1]+1
        samplesheet.cell(row=right_cell[0], column=right_cell[1]).border = Border(
            left=None,
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

    return workbook

BCLConvert_Data = [
    BCLConvert_Datum(Lane="1", Sample_ID="1041243", Index="ATCG", Index2="GCTA"),
    BCLConvert_Datum(Lane="2", Sample_ID="1414432", Index="GCTA", Index2="ATCG"),
    BCLConvert_Datum(Lane="3", Sample_ID="1414432", Index="GCTA", Index2="ATCG"),
    BCLConvert_Datum(Lane="4", Sample_ID="1414432", Index="GCTA", Index2="ATCG"),
    BCLConvert_Datum(Lane="5", Sample_ID="1414432", Index="GCTA", Index2="ATCG"),
]

DragenGermline_Data = [
    DragenGermline_Datum(Sample_ID="1041243", ReferenceGenomeDir="@ReferenceGenomeDir", VariantCallingMode="@VariantCallingMode"),
    DragenGermline_Datum(Sample_ID="1414432", ReferenceGenomeDir="@ReferenceGenomeDir", VariantCallingMode="@VariantCallingMode"),
    DragenGermline_Datum(Sample_ID="1414432", ReferenceGenomeDir="@ReferenceGenomeDir", VariantCallingMode="@VariantCallingMode"),
    DragenGermline_Datum(Sample_ID="1414432", ReferenceGenomeDir="@ReferenceGenomeDir", VariantCallingMode="@VariantCallingMode"),
    DragenGermline_Datum(Sample_ID="1414432", ReferenceGenomeDir="@ReferenceGenomeDir", VariantCallingMode="@VariantCallingMode"),
]

wb = generate_samplesheet_workbook(BCLConvert_Data, DragenGermline_Data)
wb.save("test.xlsx")